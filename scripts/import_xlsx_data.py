"""导入基准价数据和定价说明到数据库

从 knowledge/benchmark-prices/ 读取 XLSX 文件导入:
  1. 基准价2025版.xlsx → benchmark_prices 表
  2. 基准价说明.xlsx → pricing_rules 表（定价说明文本）
"""
import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

try:
    import openpyxl
except ImportError:
    print("❌ 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.benchmark_price import BenchmarkPrice

# 项目根目录
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_DIR = os.path.join(ROOT, 'knowledge', 'benchmark-prices')


def import_benchmark_prices(db: Session) -> int:
    """导入基准价2025版.xlsx"""
    xlsx_path = os.path.join(XLSX_DIR, '基准价2025版.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"❌ 文件不存在: {xlsx_path}")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            # 跳过空行和分类标题行
            seq = row[0]
            if seq is None or not isinstance(seq, (int, float)):
                continue

            internal_code = str(row[1] or '').strip()
            work_type = str(row[2] or '').strip().replace('\n', '')
            item_name = str(row[3] or '').strip()
            spec = str(row[4] or '').strip()
            unit = str(row[5] or '').strip()
            price = row[6]
            remark = str(row[7] or '').strip()

            # 跳过无价格的行
            if not unit or price is None:
                continue

            # 清理班组名称
            team_name = sheet_name.strip()
            # 去掉前面的序号如 "1、"
            if '、' in team_name:
                team_name = team_name.split('、', 1)[1]

            bp = BenchmarkPrice(
                trade_team=team_name,
                internal_code=internal_code,
                work_type=work_type,
                item_name=item_name,
                spec=spec,
                unit=unit,
                unit_price=float(price),
                remark=remark,
            )
            db.add(bp)
            count += 1
            total += 1

        print(f"  {sheet_name}: {count} 条")
        db.flush()

    wb.close()
    return total


def import_pricing_rules(db: Session) -> int:
    """导入基准价说明.xlsx（存为定价规则文本）"""
    xlsx_path = os.path.join(XLSX_DIR, '基准价说明.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"⚠️ 文件不存在: {xlsx_path}，跳过")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['说明']

    # 读取所有文本行，合并为一块
    lines = []
    for row in ws.iter_rows(values_only=True):
        if row[0]:
            lines.append(str(row[0]))

    if not lines:
        print("  ⚠️ 基准价说明为空")
        return 0

    content = '\n'.join(lines)

    # 存到 pricing_rules 表（先用原生SQL，表可能不存在时创建）
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS pricing_rules (
            id SERIAL PRIMARY KEY,
            title VARCHAR(200) NOT NULL DEFAULT '基准价说明',
            content TEXT NOT NULL,
            version VARCHAR(50) DEFAULT '2025',
            created_at TIMESTAMPTZ DEFAULT now(),
            updated_at TIMESTAMPTZ
        )
    """))
    db.execute(text("DELETE FROM pricing_rules"))
    db.execute(
        text("INSERT INTO pricing_rules (title, content, version) VALUES (:title, :content, :version)"),
        {"title": "基准价说明（2025版）", "content": content, "version": "2025"}
    )
    wb.close()
    print(f"  ✅ 定价说明导入完成（{len(lines)} 行）")
    return len(lines)


def main():
    print("=" * 50)
    print("📥 数据导入工具")
    print("=" * 50)

    # 连接数据库
    engine = create_engine(settings.DATABASE_URL)
    print(f"\n📦 连接数据库: {settings.DATABASE_URL}")

    # 确保表已创建
    from app.db.base import Base
    Base.metadata.create_all(bind=engine)
    print("✅ 数据表已就绪")

    with Session(engine) as session:
        # 1. 导入基准价
        print("\n📊 导入基准价2025版...")
        session.query(BenchmarkPrice).delete()
        session.flush()
        count = import_benchmark_prices(session)
        session.commit()
        print(f"✅ 基准价导入完成：共 {count} 条")

        # 2. 导入基准价说明
        print("\n📄 导入基准价说明...")
        rule_count = import_pricing_rules(session)
        session.commit()
        print(f"✅ 基准价说明导入完成")

    # 3. 验证
    print("\n" + "=" * 50)
    print("📋 导入结果验证")
    print("=" * 50)
    with Session(engine) as session:
        total = session.query(BenchmarkPrice).count()
        teams = session.execute(
            text("SELECT trade_team, COUNT(*) as cnt FROM benchmark_prices GROUP BY trade_team ORDER BY cnt DESC")
        ).fetchall()
        print(f"\n基准价共 {total} 条，分布：")
        for team, cnt in teams:
            print(f"  {team}: {cnt} 条")

        rules = session.execute(text("SELECT COUNT(*) FROM pricing_rules")).scalar()
        print(f"\n定价说明: {rules} 条记录")

    print("\n🎉 全部完成！")


if __name__ == '__main__':
    main()
