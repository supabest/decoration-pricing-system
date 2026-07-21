#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主材数据库 — 步骤1：批量解析供应商报价 Excel

遍历 22 个品类目录下所有 .xlsx/.xls 文件，
智能识别表头行，提取：材料名称 + 规格 + 单位 + 单价 + 品牌 + 项目来源

输出：data/seeds/main_material_raw.json（清洗前的全量数据）

用法：python3 scripts/parse_main_materials.py
"""

import os, re, json, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.expanduser('~/Downloads/主要材料和设备 (1)')
OUTPUT = os.path.join(ROOT, 'data', 'seeds', 'main_material_raw.json')

# ── 用 xlrd 读 .xls ──
def read_xls(filepath):
    """xlrd 读取 .xls，返回 sheets 列表，每个 sheet 是 [{col_index: value}]"""
    try:
        import xlrd
    except ImportError:
        print('  ⚠ xlrd not installed, skipping .xls')
        return []
    wb = xlrd.open_workbook(filepath)
    sheets = []
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        if ws.nrows < 2: continue
        rows = []
        for r in range(ws.nrows):
            row = {}
            for c in range(ws.ncols):
                val = ws.cell_value(r, c)
                if val is not None and str(val).strip():
                    row[c] = str(val).strip()
            if row:
                rows.append(row)
        if rows:
            sheets.append({'name': sn, 'rows': rows})
    return sheets

# ── 用 openpyxl 读 .xlsx ──
def read_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 2: continue
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            r = {}
            for c, val in enumerate(row):
                if val is not None and str(val).strip():
                    r[c] = str(val).strip()
            if r:
                rows.append(r)
        if rows:
            sheets.append({'name': sn, 'rows': rows})
    return sheets

# ── 智能表头检测 ──
HEADER_KEYWORDS = {
    'seq':      re.compile(r'^序号$|^项次$|^NO\.?$', re.I),
    'name':     re.compile(r'材料名称|产品名称|名称|项目内容|材料名|品名', re.I),
    'spec':     re.compile(r'规格型号|型号/规格|规格$|型号$|规格参数', re.I),
    'material': re.compile(r'材质.*（|材质$|材质\s|^材质$|材料描述', re.I),
    'thickness': re.compile(r'^厚度$|^厚度\s|材料厚度', re.I),
    'color':    re.compile(r'^颜色$|^颜色\s|表面处理|饰面颜色', re.I),
    'tech':     re.compile(r'技术参数|特征描述|项目特征描述|项目特征', re.I),
    'dim':      re.compile(r'尺寸|长宽|宽.*高.*厚|长.*宽.*高|尺寸参数', re.I),
    'unit':     re.compile(r'^单位$|计价单位|计量单位|基本计量单位', re.I),
    'price_ex_tax': re.compile(r'不含税单价|除税单价|不含税综合单价|不含税材料单价', re.I),
    'price_incl_tax': re.compile(r'含税单价|含税综合单价|含税材料单价', re.I),
    'price_mat':  re.compile(r'^主材$|材料费$|^主材费$|材料单价(?!.*不含税|.*含税)|主材单价|含税材料单价', re.I),  # 纯材料价（不含安装）
    'price':    re.compile(r'(?<!不含税)(?<!除税)(?<!含税)(?<!材料)单价|综合单价(?!.*不含税|.*含税)|信息价|不含税价格明细', re.I),
    'brand':    re.compile(r'^品牌$|厂家|供应商', re.I),
    'qty':      re.compile(r'^数量$|^工程量$|暂定数量|计价数量', re.I),
    'total':    re.compile(r'合价|总价|金额|合价.*元|总价.*元', re.I),
    'remark':   re.compile(r'^备注$|说明', re.I),
    'position': re.compile(r'^位置$|使用部位|安装位置', re.I),
}

def detect_headers(row_dict):
    """对一行 {col_index: value}，返回 {role: col_index}"""
    mapping = {}
    for ci, val in row_dict.items():
        v = str(val).replace('\n', '').strip()
        for role, pat in HEADER_KEYWORDS.items():
            if role in mapping: continue  # 已分配
            if pat.search(v):
                mapping[role] = ci
                break
    return mapping

def find_header_row(rows, max_scan=15):
    """在 sheet 的前 max_scan 行中找表头行，返回 (header_row_index, mapping)"""
    best = None
    best_score = 0
    for i, row in enumerate(rows[:max_scan]):
        # row 是 {col_index: value}
        mapping = detect_headers(row)
        score = len(mapping)
        # 必须有 name 或价格列才算有效表头
        has_price = any(k in mapping for k in ('price', 'price_ex_tax', 'price_incl_tax'))
        if ('name' in mapping or has_price) and score >= 2:
            if score > best_score:
                best_score = score
                best = (i, mapping)
    return best

# ── 价格提取辅助 ──
def parse_price(val_str):
    """从字符串中提取价格数字"""
    if not val_str: return None
    s = str(val_str).replace(',', '').replace('，', '').replace(' ', '').strip()
    # 去掉 ¥ ￥ 元 等
    s = re.sub(r'[¥￥元]', '', s)
    try:
        return round(float(s), 2)
    except ValueError:
        return None

def detect_tax_rate(rows):
    """从文件内容中检测增值税率。
    扫描前30行，查找 '13%' '9%' '3%' 等税率标识。
    返回税率小数（如0.13），默认0.13。
    """
    tax_hints = []
    for row in rows[:30]:
        for v in row.values():
            s = str(v)
            # 匹配增值税率
            m = re.search(r'增值税[专普]票?[^0-9]*(\d+)%', s)
            if m: tax_hints.append(int(m.group(1)))
            m = re.search(r'税率\s*(\d+)%', s)
            if m: tax_hints.append(int(m.group(1)))
            m = re.search(r'含税[^0-9]*\(?(\d+)%', s)
            if m: tax_hints.append(int(m.group(1)))
            m = re.search(r'含(\d+)%增值税', s)
            if m: tax_hints.append(int(m.group(1)))
    if tax_hints:
        # 取最常见的税率值
        from collections import Counter
        rate = Counter(tax_hints).most_common(1)[0][0]
        return rate / 100.0
    return 0.13  # 默认13%

def parse_unit(val_str):
    """规范化单位"""
    if not val_str: return ''
    s = str(val_str).strip()
    # 统一写法
    s = s.replace('m2', '㎡').replace('m²', '㎡').replace('m³', 'm³')
    s = s.replace('M2', '㎡').replace('m3', 'm³')
    s = s.replace('平方', '㎡').replace('平米', '㎡').replace('平方米', '㎡')
    s = s.replace('立方', 'm³').replace('立方米', 'm³')
    return s

def parse_dimensions(dim_str):
    """尝试从尺寸字符串解析 宽×高×厚（单位mm）"""
    if not dim_str: return {}
    s = str(dim_str).strip()
    result = {}
    # 匹配类似 "1320*2400*340" 或 "1020×315×40" 格式
    # 支持 * 和 × 和 x 分隔符
    m = re.match(r'[\s]*([\d.]+)\s*[\*×xX]\s*([\d.]+)\s*[\*×xX]\s*([\d.]+)', s)
    if m:
        result = {'w': float(m.group(1)), 'h': float(m.group(2)), 'd': float(m.group(3))}
    else:
        # 两段式 "2400*270"
        m2 = re.match(r'[\s]*([\d.]+)\s*[\*×xX]\s*([\d.]+)', s)
        if m2:
            w = float(m2.group(1))
            h = float(m2.group(2))
            # 启发式：数值大的可能是高，小的可能是宽
            if w > h:
                result = {'w': w, 'h': h}
            else:
                result = {'w': w, 'h': h}
    return result

def parse_spec(val_str):
    """清理规格字符串"""
    if not val_str: return ''
    s = str(val_str).replace('\n', ' ').replace('\r', ' ').strip()
    # 截断过长规格
    if len(s) > 200:
        s = s[:200] + '...'
    return s

# ── 跳过行判断 ──
SKIP_PATTERNS = [
    re.compile(r'^(合计|小计|总计|说明|备注|注：|注:|以上|以下|本表)'),
    re.compile(r'^\d+\.\d+$'),  # 纯数字序号（可能是合并单元格残留）
]

def is_skip_row(row_dict, header_map):
    """判断是否应该跳过这一行"""
    name_col = header_map.get('name')
    if name_col and name_col in row_dict:
        val = row_dict[name_col]
        for pat in SKIP_PATTERNS:
            if pat.search(val):
                return True
        # 如果 name 列内容全是数字/序号格式
        if re.match(r'^\d+[\.\、\s]*$', val.strip()):
            return True
    return False

# ── 主解析逻辑 ──
def extract_items_from_sheet(sheet, category_dir, filename, project_info):
    """从单个 sheet 提取材料记录"""
    rows = sheet['rows']
    if len(rows) < 2: return []

    result = detect_headers_and_extract(rows, category_dir, filename, project_info, sheet['name'])
    return result

def detect_headers_and_extract(rows, category_dir, filename, project_info, sheet_name):
    header_info = find_header_row(rows)
    if not header_info:
        return []

    header_idx, header_map = header_info
    tax_rate = detect_tax_rate(rows)  # 从文件检测税率
    items = []

    # 价格列：优先纯材料价 > 不含税 > 含税 > 通用单价
    price_mat_col = header_map.get('price_mat')        # 纯材料价（不含安装）
    price_ex_tax_col = header_map.get('price_ex_tax')
    price_incl_tax_col = header_map.get('price_incl_tax')
    price_col = header_map.get('price')

    name_col = header_map.get('name')
    spec_col = header_map.get('spec')
    material_col = header_map.get('material')    # 材质
    thickness_col = header_map.get('thickness')  # 厚度
    color_col = header_map.get('color')          # 颜色
    tech_col = header_map.get('tech')            # 技术参数/特征描述
    dim_col = header_map.get('dim')              # 尺寸列
    unit_col = header_map.get('unit')
    brand_col = header_map.get('brand')
    position_col = header_map.get('position')
    remark_col = header_map.get('remark')

    for i, row in enumerate(rows):
        if i <= header_idx: continue  # 跳过表头及之前行
        if not row: continue

        # 获取名称
        name = ''
        if name_col and name_col in row:
            name = row[name_col]

        # 跳过明显不是数据行的行
        if not name or len(name) > 100:
            continue
        if re.match(r'^(合计|小计|总计|说明|备注|注)', name):
            continue
        if re.match(r'^\d+[\.\、\s]*$', name.strip()):
            continue

        # 获取价格：优先纯材料价，再看不含税/含税
        price_ex_tax = None
        price_incl_tax = None

        # 第一优先：纯材料价（不含安装）
        if price_mat_col and price_mat_col in row:
            p_mat = parse_price(row[price_mat_col])
            if p_mat is not None and p_mat > 0:
                # 材料价列可能是含税也可能是不含税，根据列名判断
                col_header = ''
                for k, v in header_map.items():
                    if v == price_mat_col:
                        col_header = k
                        break
                if 'incl' in str(header_map) or '含税' in str(header_map):
                    price_incl_tax = p_mat
                else:
                    price_ex_tax = p_mat

        if price_ex_tax is None and price_incl_tax is None:
            if price_ex_tax_col and price_ex_tax_col in row:
                price_ex_tax = parse_price(row[price_ex_tax_col])
            if price_incl_tax_col and price_incl_tax_col in row:
                price_incl_tax = parse_price(row[price_incl_tax_col])

        # 如果都没取到，用通用价格列
        if price_ex_tax is None and price_incl_tax is None:
            if price_col and price_col in row:
                p = parse_price(row[price_col])
                if p is None or p <= 0:
                    continue
                # 默认当作不含税（保守）
                price_ex_tax = p
                price_incl_tax = round(p * (1 + tax_rate), 2)
            else:
                continue
        else:
            # 有至少一个价格
            if price_ex_tax is not None and price_ex_tax <= 0:
                continue
            if price_incl_tax is not None and price_incl_tax <= 0:
                continue
            if price_ex_tax is None and price_incl_tax is not None:
                if price_incl_tax <= 0: continue
                price_ex_tax = round(price_incl_tax / (1 + tax_rate), 2)
            elif price_incl_tax is None and price_ex_tax is not None:
                if price_ex_tax <= 0: continue
                price_incl_tax = round(price_ex_tax * (1 + tax_rate), 2)

        # 统一用不含税价作为主价格（用于去重）
        if price_ex_tax is None or price_ex_tax <= 0:
            continue
        price = price_ex_tax

        # 获取规格相关字段：合并 材质 + 规格型号 + 厚度 + 颜色 + 技术参数
        spec_parts = []
        if material_col and material_col in row:
            spec_parts.append(str(row[material_col]).replace('\n', ' ').strip())
        if spec_col and spec_col in row:
            spec_parts.append(str(row[spec_col]).replace('\n', ' ').strip())
        if thickness_col and thickness_col in row:
            t = str(row[thickness_col]).replace('\n', ' ').strip()
            if t and not re.match(r'^[\d.]+\s*(mm|cm|m)?$', t, re.I):
                t = '厚度' + t
            spec_parts.append(t)
        if color_col and color_col in row:
            spec_parts.append(str(row[color_col]).replace('\n', ' ').strip())
        if tech_col and tech_col in row:
            spec_parts.append(str(row[tech_col]).replace('\n', ' ').strip())
        spec = ' | '.join(p for p in spec_parts if p)

        # 截断过长规格
        if len(spec) > 300:
            spec = spec[:300] + '...'

        # 尺寸列
        dim_str = ''
        if dim_col and dim_col in row:
            dim_str = str(row[dim_col]).replace('\n', ' ').strip()
            if dim_str and dim_str not in spec:
                spec = (spec + '  【尺寸】' + dim_str).strip()

        unit = parse_unit(row[unit_col]) if unit_col and unit_col in row else ''
        brand = row[brand_col] if brand_col and brand_col in row else ''
        position = row[position_col] if position_col and position_col in row else ''
        remark = row[remark_col] if remark_col and remark_col in row else ''

        # 解析尺寸：尝试从 dim_str 提取 长/宽/高
        parsed_dim = parse_dimensions(dim_str)

        items.append({
            'category_dir': category_dir,
            'category_name': CATEGORY_NAMES.get(category_dir, category_dir),
            'material_name': name,
            'spec': spec,
            'unit': unit,
            'price': price,                     # 不含税单价（主价格，用于去重和排序）
            'price_incl_tax': price_incl_tax,   # 含税单价
            'tax_rate': tax_rate,               # 增值税率
            'brand': brand,
            'position': position,
            'remark': remark,
            'source_file': filename,
            'project': project_info.get('project', ''),
            'project_type': project_info.get('type', ''),
            'sheet': sheet_name,
            'dim_raw': dim_str,             # 原始尺寸字符串
            'dim_width': parsed_dim.get('w'),   # 宽 mm
            'dim_height': parsed_dim.get('h'),  # 高 mm
            'dim_depth': parsed_dim.get('d'),   # 厚/深 mm
        })

    return items

# ── 品类目录名映射 ──
CATEGORY_NAMES = {
    '101瓷砖': '瓷砖',
    '102石材': '石材',
    '103木地板': '木地板',
    '104玻璃': '玻璃',
    '105不锈钢': '不锈钢',
    '106铝板': '铝板',
    '107地毯': '地毯',
    '108织物': '织物/窗帘',
    '109油漆涂料': '油漆涂料',
    '110金属门窗': '金属门窗',
    '111卫浴五金': '卫浴五金',
    '112智能设备': '智能设备',
    '113暖通设备': '暖通设备',
    '114照明设备': '照明设备',
    '115墙纸皮革': '墙纸皮革',
    '117厨房设备': '厨房设备',
    '118石膏制品': '石膏制品',
    '119木饰面及木门柜体': '木饰面/木门/柜体',
    '120塑胶地板': '塑胶地板',
    '121亚克力': '亚克力',
    '122成品隔断': '成品隔断',
}

# ── 从文件名提取项目信息 ──
def extract_project_info(filename):
    """从文件名解析项目和类型"""
    info = {'project': '', 'type': ''}
    s = filename

    # 提取年份
    year_m = re.search(r'(20\d{2})_', s)

    # 提取项目名
    proj_patterns = [
        (r'聚龙湾', '聚龙湾'),
        (r'大干围', '大干围'),
        (r'珑璟台', '珑璟台'),
        (r'坑口', '坑口'),
        (r'柏悦湾', '柏悦湾'),
        (r'先烈路', '先烈路'),
        (r'东方明珠', '东方明珠'),
        (r'三一云湖', '三一云湖'),
        (r'珠江', '珠江'),
    ]
    for pat, name in proj_patterns:
        if pat in s:
            info['project'] = name
            break

    # 提取项目类型
    type_patterns = [
        (r'样板房', '样板房'),
        (r'批量住宅', '批量住宅'),
        (r'售楼部', '售楼部'),
        (r'展示区', '展示区'),
        (r'公区', '公区'),
        (r'会所', '会所'),
        (r'幼儿园', '幼儿园'),
        (r'展厅', '展厅'),
        (r'招商中心', '招商中心'),
        (r'办公室', '办公室'),
    ]
    for pat, tname in type_patterns:
        if pat in s:
            info['type'] = tname
            break

    return info

# ── 主流程 ──
def main():
    if not os.path.isdir(DATA_DIR):
        print(f'❌ 数据目录不存在: {DATA_DIR}')
        sys.exit(1)

    all_items = []
    stats = {'files': 0, 'sheets': 0, 'items': 0, 'skipped_dirs': [], 'errors': []}

    for cat_dir in sorted(os.listdir(DATA_DIR)):
        if cat_dir.startswith('.'): continue
        cat_path = os.path.join(DATA_DIR, cat_dir)
        if not os.path.isdir(cat_path): continue

        files = [f for f in os.listdir(cat_path)
                 if not f.startswith('~') and not f.startswith('.')
                 and (f.endswith('.xlsx') or f.endswith('.xls'))]

        if not files:
            stats['skipped_dirs'].append(cat_dir)
            continue

        print(f'\n📁 {cat_dir} ({len(files)} files)')

        for f in sorted(files):
            # 跳过标记为"仅供格式参考"的文件
            if '仅供格式参考' in f or '不作价格参考' in f:
                print(f'  ⏭ {f}: 仅供格式参考，跳过')
                continue
            # 跳过石材护理/精保洁报价（服务费而非材料费）
            if cat_dir == '102石材' and ('护理' in f or '精保洁' in f or '保洁' in f):
                print(f'  ⏭ {f}: 石材护理（服务费），跳过')
                continue
            fp = os.path.join(cat_path, f)
            proj_info = extract_project_info(f)

            try:
                if f.endswith('.xls'):
                    sheets = read_xls(fp)
                else:
                    sheets = read_xlsx(fp)

                if not sheets:
                    continue

                stats['files'] += 1
                file_items = 0

                for sheet in sheets:
                    items = detect_headers_and_extract(
                        sheet['rows'], cat_dir, f, proj_info, sheet['name']
                    )
                    if items:
                        all_items.extend(items)
                        file_items += len(items)
                        stats['items'] += len(items)

                stats['sheets'] += len(sheets)
                print(f'  ✓ {f}: {file_items} items ({len(sheets)} sheets)')

            except Exception as e:
                msg = f'{cat_dir}/{f}: {e}'
                print(f'  ❌ {msg[:100]}')
                stats['errors'].append(msg)

    # ── 输出 ──
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(all_items, f, ensure_ascii=False, indent=2)

    print(f'\n{"="*60}')
    print(f'✅ 解析完成')
    print(f'  品类目录: {len([d for d in os.listdir(DATA_DIR) if not d.startswith(".")])} 个')
    print(f'  有效文件: {stats["files"]} 个')
    print(f'  Sheets:   {stats["sheets"]} 个')
    print(f'  提取记录: {stats["items"]} 条')
    print(f'  跳过目录: {stats["skipped_dirs"]}')
    print(f'  错误:     {len(stats["errors"])} 个')
    print(f'  输出:     {OUTPUT}')
    if stats['errors']:
        print(f'\n  错误详情（前10）:')
        for e in stats['errors'][:10]:
            print(f'    - {e[:120]}')

    # ── 摘要统计 ──
    if all_items:
        cats = {}
        for it in all_items:
            c = it['category_name']
            cats[c] = cats.get(c, 0) + 1
        print(f'\n  按品类分布:')
        for c, n in sorted(cats.items(), key=lambda x: -x[1]):
            print(f'    {c}: {n} 条')

        prices = [it['price'] for it in all_items]
        print(f'\n  价格范围: {min(prices):.2f} ~ {max(prices):.2f} 元')
        print(f'  有单位记录: {sum(1 for it in all_items if it["unit"])} 条')
        print(f'  有规格记录: {sum(1 for it in all_items if it["spec"])} 条')
        print(f'  有品牌记录: {sum(1 for it in all_items if it["brand"])} 条')

if __name__ == '__main__':
    main()
