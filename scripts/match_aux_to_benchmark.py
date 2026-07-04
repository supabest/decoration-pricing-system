#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将《基准价2024（格式化）.xlsx》中的辅材费匹配到 497 条人工基准价。
辅材费计算（按清单序号分组）：
    aux_total = Σ(组内 K) + Σ(组内 H×J，仅 H 有值的行)
输出：data/seeds/aux_matching_review.xlsx 供人工审核。
"""
import json, re, os
from collections import defaultdict
import openpyxl
from rapidfuzz import fuzz

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = '/Users/alick/Downloads/基准价2024（格式化）.xlsx'
BENCH = os.path.join(ROOT, 'data/seeds/benchmark_items_supabase.json')
OUT_XLSX = os.path.join(ROOT, 'data/seeds/aux_matching_review.xlsx')
OUT_JSON = os.path.join(ROOT, 'data/seeds/aux_excel_items.json')

# ---------- 1. 解析 Excel ----------
def parse_sheet(ws, sheet_label):
    """按序号(ZS开头)分组，返回结构化项目列表"""
    items = []
    cur = None
    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value        # 序号
        b = ws.cell(r, 2).value        # 项目名称
        c = ws.cell(r, 3).value        # 项目特征
        d = ws.cell(r, 4).value        # 单位
        e = ws.cell(r, 5).value        # 工种
        g = ws.cell(r, 7).value        # 主材名称
        h = ws.cell(r, 8).value        # 主材单价 H
        i = ws.cell(r, 9).value        # 单位
        j = ws.cell(r, 10).value       # 消耗量 J
        k = ws.cell(r, 11).value       # 辅材费 K
        if a and str(a).startswith('ZS'):
            # 新清单项目
            cur = {
                'source_sheet': sheet_label,
                'code': str(a),
                'name': _clean(b),
                'spec': _clean(c),
                'unit': _clean(d),
                'work_type': _clean(e),
                'rate_mats': [],   # [{name, price, consume}]
                'fixed_k': 0.0,
                'row_start': r,
            }
            items.append(cur)
        if cur is None:
            continue
        # 累加 K
        if isinstance(k, (int, float)):
            cur['fixed_k'] += float(k)
        # 累加 RATE 材料 (H 有值)
        if isinstance(h, (int, float)) and h > 0:
            cur['rate_mats'].append({
                'name': _clean(g),
                'price': float(h),
                'consume': float(j) if isinstance(j, (int, float)) else 0.0,
                'unit': _clean(i),
            })
    # 计算合计
    for it in items:
        rate_sum = sum(m['price'] * m['consume'] for m in it['rate_mats'])
        it['rate_sum'] = round(rate_sum, 4)
        it['fixed_k'] = round(it['fixed_k'], 4)
        it['aux_total'] = round(it['fixed_k'] + rate_sum, 4)
        it['rule_type'] = 'GROUP' if it['rate_mats'] else 'FIXED'
    return items

def _clean(v):
    if v is None: return ''
    return re.sub(r'\s+', ' ', str(v)).strip()

wb = openpyxl.load_workbook(XLSX, data_only=True)
excel_items = parse_sheet(wb['ZS02-A-精装修工程'], 'A精装修') \
            + parse_sheet(wb['ZS02-B- 精装修水电'], 'B水电')
print(f'Excel 解析：共 {len(excel_items)} 个清单项目')
print(f'  GROUP 型(有RATE子行): {sum(1 for i in excel_items if i["rate_mats"])}')
print(f'  FIXED 型: {sum(1 for i in excel_items if not i["rate_mats"])}')

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(excel_items, f, ensure_ascii=False, indent=2)

# ---------- 2. 加载基准价 ----------
with open(BENCH, encoding='utf-8') as f:
    bj = json.load(f)
bench_items = bj if isinstance(bj, list) else list(bj.values())[0]
print(f'基准价：共 {len(bench_items)} 条')

# ---------- 3. 语义匹配 ----------
# 同义词归一化，减少命名差异导致的错配
SYN = {
    '地砖': '瓷砖', '墙砖': '瓷砖', '面砖': '瓷砖', '玻化砖': '瓷砖',
    '天棚': '吊顶', '天花': '吊顶', '顶面': '吊顶',
    '腻子': '批灰', '涂料': '乳胶漆', '墙面漆': '乳胶漆',
    '墙纸': '壁纸', '墙布': '壁纸',
    '木地板': '木地板', '实木地板': '木地板',
    '水泥沙': '水泥砂浆', '水泥沙浆': '水泥砂浆',
    '镶贴': '铺贴', '粘贴': '铺贴',
    '墙身': '墙面', '墙柱面': '墙面', '墙柱': '墙面',
    '楼地面': '地面',
}
# 工种映射（基准价工种 → Excel 工种集合）。命中加分，不命中减分
WT_MAP = {
    '泥水工': {'抹灰镶贴工', '砌筑工', '防水工'},
    '防水工': {'防水工', '抹灰镶贴工'},
    '木工': {'装饰木工'},
    '油漆工': {'油漆工'},
    '铁工': {'金属制品安装工'},
    '水电工': {'管工', '电工'},
    '玻璃工': set(),
    '幕墙工': set(),
    '拆除班组': set(),
    '专业班组': set(),
}
def normalize(s):
    s = _clean(s)
    s = re.sub(r'[（(].*?[）)]', '', s)      # 去括号注释
    s = re.sub(r'[\-—－·/、,，。.;；:：\s]+', '', s)
    for k, v in SYN.items():
        s = s.replace(k, v)
    return s

def suffix_after_dash(s):
    """取最后一个破折号后的具体项目名（去掉类别前缀）"""
    for sep in ['-', '—', '－']:
        if sep in s:
            return s.rsplit(sep, 1)[-1]
    return s

def match_score(bench, excel):
    """返回 0-100 综合语义相似度"""
    bn_full = normalize(bench['item_name'])
    bn_suf = normalize(suffix_after_dash(bench['item_name']))
    en = normalize(excel['name'])
    # 名称：全名 + 后缀 三路取最大（后缀对"类别-项目"格式最关键）
    s1 = fuzz.token_set_ratio(bn_full, en)
    s2 = fuzz.token_set_ratio(bn_suf, en)
    s3 = fuzz.partial_ratio(bn_suf, en) if min(len(bn_suf), len(en)) >= 3 else 0
    name_score = max(s1, s2, s3)
    # 项目特征关键词重叠（材料/工艺术语）
    bs = set(re.findall(r'[一-龥]{2,}', normalize(bench.get('spec', ''))))
    es = set(re.findall(r'[一-龥]{2,}', normalize(excel.get('spec', ''))))
    spec_score = (len(bs & es) / max(1, min(len(bs), len(es))) * 100) if (bs and es) else 0
    # 工种映射加分/减分
    bwt = _clean(bench.get('work_type', ''))
    ewt = _clean(excel.get('work_type', ''))
    wt_bonus = 0
    if bwt in WT_MAP:
        if ewt in WT_MAP[bwt]:
            wt_bonus = 12
        elif WT_MAP[bwt]:  # Excel 有对应工种但本行不是
            wt_bonus = -8
    return min(100, round(name_score * 0.80 + spec_score * 0.12 + wt_bonus))

results = []
NO_EXCEL_TEAMS = {'幕墙工', '玻璃工', '拆除班组', '专业班组', '架子工\n（专业分包）'}
for bi in bench_items:
    bwt = _clean(bi.get('work_type', ''))
    # Excel 无此类工种 → 直接标记无数据，不强行匹配
    if bwt in NO_EXCEL_TEAMS or (bwt in WT_MAP and not WT_MAP[bwt]):
        results.append({
            'bench_team': bi.get('trade_team', ''), 'bench_work': bwt,
            'bench_name': bi.get('item_name', ''), 'bench_unit': bi.get('unit', ''),
            'bench_labor': bi.get('unit_price', ''),
            'matched_code': '—', 'matched_name': '（Excel无此类项目）', 'matched_sheet': '—',
            'score': 0, 'second_score': 0, 'status': 'no_excel',
            'rule_type': '—', 'rate_detail': '', 'fixed_k': 0, 'rate_sum': 0, 'aux_total': 0,
            'bench_spec': (bi.get('spec', '') or '')[:60],
        })
        continue
    scored = [(match_score(bi, ei), ei) for ei in excel_items]
    scored.sort(key=lambda x: x[0], reverse=True)
    best_s, best = scored[0]
    second_s = scored[1][0] if len(scored) > 1 else 0
    # 状态判定
    if best_s >= 82:
        status = 'matched'
    elif best_s >= 68 and (best_s - second_s) >= 8:
        status = 'matched'   # 明显优于次选
    elif best_s >= 58:
        status = 'ambiguous'
    else:
        status = 'unmatched'
    rate_desc = ' + '.join(f"{m['name']}({m['price']}×{m['consume']})" for m in best['rate_mats'])
    results.append({
        'bench_team': bi.get('trade_team', ''),
        'bench_work': bi.get('work_type', ''),
        'bench_name': bi.get('item_name', ''),
        'bench_unit': bi.get('unit', ''),
        'bench_labor': bi.get('unit_price', ''),
        'matched_code': best['code'],
        'matched_name': best['name'],
        'matched_sheet': best['source_sheet'],
        'score': best_s,
        'second_score': second_s,
        'status': status,
        'rule_type': best['rule_type'],
        'rate_detail': rate_desc,
        'fixed_k': best['fixed_k'],
        'rate_sum': best['rate_sum'],
        'aux_total': best['aux_total'],
        'bench_spec': (bi.get('spec', '') or '')[:60],
    })

# ---------- 4. 输出审核 Excel ----------
from openpyxl.styles import Font, PatternFill, Alignment
out = openpyxl.Workbook()
ws = out.active
ws.title = '辅材匹配审核'
headers = ['班组','工种','基准价项目名称','单位','人工单价','匹配状态','匹配置信度',
           '次选分差','命中Excel编码','命中Excel项目','来源薄','规则类型',
           '辅材费合计','固定K','RATE明细','RATE小计','基准价项目特征(前60字)']
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

fill_red = PatternFill('solid', fgColor='FFC7CE')
fill_yel = PatternFill('solid', fgColor='FFEB9C')
fill_grn = PatternFill('solid', fgColor='C6EFCE')
fill_gry = PatternFill('solid', fgColor='D9D9D9')
for r in results:
    row = [r['bench_team'], r['bench_work'], r['bench_name'], r['bench_unit'], r['bench_labor'],
           r['status'], r['score'], r['score']-r['second_score'], r['matched_code'], r['matched_name'],
           r['matched_sheet'], r['rule_type'], r['aux_total'], r['fixed_k'], r['rate_detail'], r['rate_sum'],
           r['bench_spec']]
    ws.append(row)
    cell_status = ws.cell(ws.max_row, 6)
    cell_status.fill = {'matched': fill_grn, 'ambiguous': fill_yel,
                        'unmatched': fill_red, 'no_excel': fill_gry}[r['status']]

# 列宽
widths = [10,8,28,6,8,10,10,8,14,28,10,10,10,8,40,10,40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# 汇总 sheet
ws2 = out.create_sheet('汇总')
from collections import Counter
sc = Counter(r['status'] for r in results)
ws2.append(['状态','数量','占比'])
for k in ['matched','ambiguous','unmatched','no_excel']:
    ws2.append([k, sc.get(k,0), f"{sc.get(k,0)/len(results)*100:.1f}%"])
ws2.append([])
ws2.append(['说明：'])
ws2.append(['matched(绿) = 高置信匹配，可直接采用'])
ws2.append(['ambiguous(黄) = 有候选但不够确定，需人工复核'])
ws2.append(['unmatched(红) = 有 Excel 数据但未匹配上，需手动指定'])
ws2.append(['no_excel(灰) = Excel 无此类工种项目(幕墙/玻璃/拆除等)，辅材费=0'])

out.save(OUT_XLSX)
print(f'\n已输出审核文件: {OUT_XLSX}')
print(f'匹配结果: matched={sc.get("matched",0)}  ambiguous={sc.get("ambiguous",0)}  unmatched={sc.get("unmatched",0)}  (共{len(results)})')
