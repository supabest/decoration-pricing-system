"""Excel 导入导出工具"""
from openpyxl import Workbook, load_workbook
from typing import List, Dict


def import_from_excel(file_path: str) -> List[Dict]:
    """从 Excel 导入数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data


def export_to_excel(headers: List[str], rows: List[Dict], file_path: str):
    """导出数据到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(file_path)
